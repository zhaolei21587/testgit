import xlrd
import logging
from openpyxl import load_workbook
import pandas as pd
import xlwings


def read_excel(filename, sheetname):
    '''
    读取Excel数据返回一个包含字典的列表
    :param filename: excel文件的名称
    :return: 一个包含字典的列表和sheet页中的行数
    '''
    tabletitle = []
    result_list = []
    filepath = r'E:\job\Project\SMSAPI\Data\{}'.format(filename)
    table = xlrd.open_workbook(filepath)
    sheet = table.sheet_by_name(sheetname)
    rows = sheet.nrows
    column = sheet.ncols
    for i in range(1,rows):
        rowvalue = sheet.row_values(i)
        for j in range(column):
            key = sheet.cell_value(0, j)
            tabletitle.append(key)
        result_list.append(dict(zip(tabletitle,rowvalue)))
    return result_list,rows

def write_excel(filename, rows, cols, values):
    '''
    写数据到Excel指定的行和列中
    :param filename: Excel文件名称
    :param rows: 准备写入数据的行数
    :param cols: 准备写入数据的列数
    :param values: 写入的内容
    :return: 无
    '''
    path = r'.\Data\{}'.format(filename)
    wb = load_workbook(path)
    sheet = wb.active
    sheet.cell(row=rows,column=cols).value = values
    wb.save(path)

def drop_excel(filename, m_col, m_text):
    '''
    该方法有问题无法使用
    :param filename:
    :param m_col:
    :param m_text:
    :return:
    '''
    filepath = r'E:\job\Project\SMSAPI\Data\{}'.format(filename)
    app = xlwings.App(visible=False)
    wb = app.books.open(filepath)
    sheet = wb.sheets[0]
    rows = sheet.used_range.last_cell.row
    for i in range(rows,1,-1):
        if sheet.range(m_col+str(i)).value == m_text:
            print('对第'+ str(i) + "行进行了删除行操作")
            sheet.range(m_col+str(i)).api.EntireRow.Delete()
            k = i + 1
    wb.close()





if __name__ == '__main__':
    drop_excel('testlink_result.xlsx','2','p')
